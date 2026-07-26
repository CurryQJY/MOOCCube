from __future__ import annotations

import importlib.util
from pathlib import Path

import openpyxl
import pandas as pd
import pytest


SCRIPT = (
    Path(__file__).resolve().parents[1]
    / "paper_aaai27/scripts/export_overall_baseline_comparison.py"
)
SPEC = importlib.util.spec_from_file_location(
    "export_overall_baseline_comparison", SCRIPT
)
assert SPEC is not None and SPEC.loader is not None
MODULE = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(MODULE)


def test_weighted_overall_uses_course_counts() -> None:
    value = MODULE.weighted_overall(0.2, 2, 0.8, 8)

    assert value == pytest.approx(0.68)
    assert value != pytest.approx(0.5)


def test_validate_direct_overall_rejects_mismatch() -> None:
    with pytest.raises(ValueError, match="direct overall mismatch"):
        MODULE.validate_direct_overall(0.2, 0.25, tolerance=1e-8)


def test_unavailable_row_contains_no_numeric_metrics() -> None:
    row = MODULE.unavailable_row(
        "MOOCCube", "PCGNN", 2025, "missing warm targets"
    )

    assert row["status"] == "unavailable_missing_warm_targets"
    assert all(pd.isna(row[metric]) for metric in MODULE.METRICS)


def test_build_seed_row_reconstructs_all_metrics() -> None:
    row = MODULE.build_seed_row(
        dataset="Toy",
        method="Baseline",
        seed=2025,
        cold={metric: 0.2 for metric in MODULE.METRICS},
        hot={metric: 0.8 for metric in MODULE.METRICS},
        cold_count=2,
        hot_count=8,
        cold_source="cold.csv",
        hot_source="hot.csv",
        direct={metric: 0.68 for metric in MODULE.METRICS},
    )

    assert row["status"] == "ready"
    assert row["aggregation_route"] == "cold_hot_course_count_weighted"
    assert row["R@10"] == pytest.approx(0.68)
    assert row["cold_count"] == 2
    assert row["hot_count"] == 8


def test_build_seed_row_rejects_missing_metric() -> None:
    cold = {metric: 0.2 for metric in MODULE.METRICS if metric != "N@20"}

    with pytest.raises(ValueError, match="missing metric N@20"):
        MODULE.build_seed_row(
            dataset="Toy",
            method="Baseline",
            seed=2025,
            cold=cold,
            hot={metric: 0.8 for metric in MODULE.METRICS},
            cold_count=2,
            hot_count=8,
            cold_source="cold.csv",
            hot_source="hot.csv",
            direct=None,
        )


def test_collect_seed_rows_covers_current_main_table_artifacts() -> None:
    detail, coverage = MODULE.collect_seed_rows()

    assert len(coverage) == 108
    assert len(detail) == 108
    assert set(coverage.status) == {"ready"}
    assert set(detail.method) == set(MODULE.METHOD_ORDER)
    assert set(detail.seed) == {2025, 2026, 2027}
    assert set(detail.dataset) == {"MOOCCube", "Junyi", "COCO"}
    assert (
        detail.loc[detail.method == "KGRec", "aggregation_route"]
        == "direct_all_item_macro_validated"
    ).all()
    pcgnn = detail.loc[detail.method == "PCGNN"]
    assert len(pcgnn) == 9
    assert (
        pcgnn["aggregation_route"] == "pcgnn_cold_target_only_report"
    ).all()
    assert (pcgnn["hot_count"] == 0).all()
    assert pcgnn["reason"].str.contains("no warm-target PCGNN metrics").all()


def test_mooccube_aldi_uses_clean_gpu_rerun_artifacts() -> None:
    detail, _ = MODULE.collect_seed_rows()

    aldi = detail[
        (detail.dataset == "MOOCCube")
        & (detail.method == "ALDI")
        & (detail.seed == 2025)
    ].iloc[0]

    assert "outputs\\score_parity\\mooccube_seed2025" in aldi["cold_source"]
    assert "outputs\\score_parity\\mooccube_seed2025" in aldi["hot_source"]
    assert "content_delta_pop5" not in aldi["hot_source"]
    assert aldi["R@10"] == pytest.approx(0.16399689943388226)
    assert aldi["N@10"] == pytest.approx(0.10620417854331396)


def test_summarize_adds_sample_std_rank_and_ckg_improvement() -> None:
    rows = []
    for seed, ours_value in zip((2025, 2026, 2027), (0.21, 0.22, 0.23)):
        rows.append(
            {
                "dataset": "Toy",
                "method": "Baseline",
                "seed": seed,
                "status": "ready",
                **{metric: 0.2 for metric in MODULE.METRICS},
            }
        )
        rows.append(
            {
                "dataset": "Toy",
                "method": "CKG-RL",
                "seed": seed,
                "status": "ready",
                **{metric: ours_value for metric in MODULE.METRICS},
            }
        )

    summary = MODULE.summarize_ready(pd.DataFrame(rows))
    ours = summary[
        (summary.dataset == "Toy") & (summary.method == "CKG-RL")
    ].iloc[0]

    assert ours["R@10_std"] == pytest.approx(0.01)
    assert ours["R@10_rank"] == 1
    assert ours["R@10_strongest_baseline"] == pytest.approx(0.2)
    assert ours["R@10_relative_improvement"] == pytest.approx(0.1)


def test_summarize_rejects_missing_seed() -> None:
    detail = pd.DataFrame(
        [
            {
                "dataset": "Toy",
                "method": "Baseline",
                "seed": seed,
                "status": "ready",
                **{metric: 0.2 for metric in MODULE.METRICS},
            }
            for seed in (2025, 2026)
        ]
    )

    with pytest.raises(ValueError, match="requires seeds 2025, 2026, 2027"):
        MODULE.summarize_ready(detail)


def test_build_wide_uses_main_table_metrics() -> None:
    detail, _ = MODULE.collect_seed_rows()
    summary = MODULE.summarize_ready(detail)

    wide = MODULE.build_wide(summary)

    assert list(wide.method) == list(MODULE.METHOD_ORDER)
    assert "MOOCCube_R@5" in wide.columns
    assert "COCO_N@10" in wide.columns
    assert "MOOCCube_R@20" not in wide.columns


def test_write_outputs_creates_required_workbook_and_csvs(
    tmp_path: Path,
) -> None:
    detail, coverage = MODULE.collect_seed_rows()
    summary = MODULE.summarize_ready(detail)

    paths = MODULE.write_outputs(detail, summary, coverage, tmp_path)

    assert {path.name for path in paths} == {
        "overall_baseline_comparison.xlsx",
        "overall_main_table.xlsx",
        "overall_baseline_summary.csv",
        "overall_baseline_seed_detail.csv",
        "overall_baseline_wide.csv",
        "overall_baseline_coverage.csv",
        "overall_main_table.tex",
        "overall_main_table_aaai.tex",
    }
    workbook = openpyxl.load_workbook(
        tmp_path / "overall_baseline_comparison.xlsx", data_only=False
    )
    assert workbook.sheetnames == ["Summary", "Seed_Detail", "Coverage"]
    assert workbook["Summary"].freeze_panes == "A2"
    assert workbook["Seed_Detail"].auto_filter.ref is not None
    assert workbook["Coverage"]["A1"].font.name == "Arial"

    csv_summary = pd.read_csv(tmp_path / "overall_baseline_summary.csv")
    assert len(csv_summary) == 36
    assert set(csv_summary.method) == set(MODULE.METHOD_ORDER)

    main_table = openpyxl.load_workbook(
        tmp_path / "overall_main_table.xlsx", data_only=False
    )
    assert main_table.sheetnames == ["Overall_Main_Table"]
    sheet = main_table["Overall_Main_Table"]
    assert {str(cell_range) for cell_range in sheet.merged_cells.ranges} >= {
        "A1:A2",
        "B1:E1",
        "F1:I1",
        "J1:M1",
    }
    assert [sheet.cell(1, column).value for column in (2, 6, 10)] == [
        "MOOCCube",
        "Junyi",
        "COCO",
    ]
    assert [sheet.cell(2, column).value for column in range(2, 6)] == [
        "R@5",
        "R@10",
        "N@5",
        "N@10",
    ]
    assert [sheet.cell(row, 1).value for row in range(3, 15)] == list(
        MODULE.METHOD_ORDER
    )
    assert sheet["B11"].value == pytest.approx(0.0240777247409046)
    assert sheet["A15"].value == "Imp."
    assert sheet["B15"].number_format == "+0.0%;-0.0%;0.0%"
    assert sheet["C13"].font.bold is True
    assert sheet["C10"].font.underline == "single"
    assert sheet["A1"].font.name == "Arial"

    table_tex = (tmp_path / "overall_main_table.tex").read_text(encoding="ascii")
    assert "Overall Item-Macro performance" in table_tex
    assert (
        r"PCGNN\methodref{TKDD'24} & 0.0241 & 0.0520 & 0.0145 & 0.0234"
        in table_tex
    )
    assert r"\emph{Imp.}" in table_tex

    table_fragment = (
        tmp_path / "overall_main_table_aaai.tex"
    ).read_text(encoding="ascii")
    assert r"\begin{table*}[t]" in table_fragment
    assert r"\documentclass" not in table_fragment
    assert (
        r"PCGNN\methodref{TKDD'24} & 0.0241 & 0.0520 & 0.0145 & 0.0234"
        in table_fragment
    )
