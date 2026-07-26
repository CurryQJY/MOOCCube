from __future__ import annotations

import json
import math
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
PAPER = ROOT / "paper_aaai27"
SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

import audit_significance_inputs as cold_audit
import export_warm_target_table as warm_export


METRICS = ("R@5", "R@10", "R@20", "N@5", "N@10", "N@20")
MAIN_METRICS = ("R@5", "R@10", "N@5", "N@10")
DATASETS = ("MOOCCube", "Junyi", "COCO")
SEEDS = (2025, 2026, 2027)
METHOD_ORDER = (
    "Popularity",
    "BPR",
    "DropoutNet",
    "LightGCN",
    "CCFCRec",
    "ALDI",
    "KGRec",
    "CGRC",
    "PCGNN",
    "USIM",
    "SEMCo",
    "CKG-RL",
)
OUTPUT_DIR = PAPER / "figures/overall_baseline_comparison"
METHOD_LABELS = {
    "Popularity": "Popularity",
    "BPR": r"BPR\methodref{UAI'09}",
    "DropoutNet": r"DropoutNet\methodref{NeurIPS'17}",
    "LightGCN": r"LightGCN\methodref{SIGIR'20}",
    "CCFCRec": r"CCFCRec\methodref{WWW'23}",
    "ALDI": r"ALDI\methodref{SIGIR'23}",
    "KGRec": r"KGRec\methodref{KDD'23}",
    "CGRC": r"CGRC\methodref{SIGIR'24}",
    "PCGNN": r"PCGNN\methodref{TKDD'24}",
    "USIM": r"USIM\methodref{NeurIPS'24}",
    "SEMCo": r"SEMCo\methodref{SIGIR'26}",
    "CKG-RL": r"\textbf{CKG-RL}",
}
OVERALL_MAIN_TABLE_CAPTION = (
    "Overall Item-Macro performance across cold and warm target courses under "
    "the strict course-cold split. Values are three-seed means reconstructed "
    "per seed by weighting cold and hot course-macro metrics by their "
    "evaluated target-course counts. Bold and underlined values denote the "
    "best and second-best available means. Each method uses its native "
    "ranking score under the common splits and evaluator. PCGNN uses retained "
    "cold-target strict reports from the formal KG-warm runs because no "
    "warm-target PCGNN metrics were retained. Imp.: CKG-RL relative change "
    "against the strongest available baseline; this descriptive overall table "
    "does not add significance markers."
)
VENUE_TAGS = (
    "UAI'09=BPR; NeurIPS'17=DropoutNet; SIGIR'20=LightGCN; "
    "WWW'23=CCFCRec; SIGIR'23=ALDI; KDD'23=KGRec; SIGIR'24=CGRC; "
    "TKDD'24=PCGNN; NeurIPS'24=USIM; SIGIR'26=SEMCo."
)


def weighted_overall(
    cold: float,
    cold_count: int,
    hot: float,
    hot_count: int,
) -> float:
    if not math.isfinite(float(cold)) or not math.isfinite(float(hot)):
        raise ValueError("cold and hot metrics must be finite")
    if cold_count <= 0 or hot_count <= 0:
        raise ValueError("cold and hot course counts must be positive")
    return (
        float(cold) * int(cold_count) + float(hot) * int(hot_count)
    ) / (int(cold_count) + int(hot_count))


def validate_direct_overall(
    reconstructed: float,
    direct: float,
    tolerance: float = 5e-5,
) -> None:
    if not math.isfinite(float(direct)) or abs(
        float(reconstructed) - float(direct)
    ) > float(tolerance):
        raise ValueError(
            "direct overall mismatch: "
            f"reconstructed={reconstructed:.12g}, direct={direct:.12g}"
        )


def unavailable_row(
    dataset: str,
    method: str,
    seed: int,
    reason: str,
) -> dict[str, object]:
    return {
        "dataset": dataset,
        "method": method,
        "seed": int(seed),
        "status": "unavailable_missing_warm_targets",
        "reason": reason,
        "aggregation_route": "unavailable",
        "cold_count": 0,
        "hot_count": 0,
        "cold_source": "",
        "hot_source": "",
        **{metric: math.nan for metric in METRICS},
    }


def build_seed_row(
    dataset: str,
    method: str,
    seed: int,
    cold: dict[str, float],
    hot: dict[str, float],
    cold_count: int,
    hot_count: int,
    cold_source: str,
    hot_source: str,
    direct: dict[str, float] | None,
) -> dict[str, object]:
    row: dict[str, object] = {
        "dataset": dataset,
        "method": method,
        "seed": int(seed),
        "status": "ready",
        "reason": "",
        "aggregation_route": "cold_hot_course_count_weighted",
        "cold_count": int(cold_count),
        "hot_count": int(hot_count),
        "cold_source": cold_source,
        "hot_source": hot_source,
    }
    for metric in METRICS:
        if metric not in cold or metric not in hot:
            raise ValueError(f"missing metric {metric} for {dataset}/{method}/{seed}")
        value = weighted_overall(
            cold[metric], cold_count, hot[metric], hot_count
        )
        if direct is not None and metric in direct:
            validate_direct_overall(value, direct[metric])
        row[metric] = value
    return row


def first_existing(paths: tuple[Path, ...]) -> Path | None:
    return next((path for path in paths if path.exists()), None)


def relative(path: Path | None) -> str:
    if path is None:
        return ""
    try:
        return str(path.relative_to(ROOT))
    except ValueError:
        return str(path)


def per_item_metrics(path: Path) -> tuple[dict[str, float], int]:
    frame = pd.read_csv(path)
    if frame.empty:
        raise ValueError(f"empty per-item file: {path}")
    missing = [metric for metric in METRICS if metric not in frame.columns]
    if missing:
        raise ValueError(f"missing per-item metrics {missing}: {path}")
    metrics = {
        metric: float(pd.to_numeric(frame[metric], errors="raise").mean())
        for metric in METRICS
    }
    return metrics, int(len(frame))


def load_result_payload(path: Path) -> dict[str, object]:
    if path.suffix.lower() == ".json":
        payload = json.loads(path.read_text(encoding="utf-8-sig"))
        if isinstance(payload, list):
            return payload[0] if payload else {}
        return payload if isinstance(payload, dict) else {}
    if path.suffix.lower() == ".csv":
        frame = pd.read_csv(path)
        return frame.iloc[0].to_dict() if not frame.empty else {}
    raise ValueError(f"unsupported result source: {path}")


def result_group_metrics(
    path: Path,
    group: str,
) -> tuple[dict[str, float], int]:
    payload = load_result_payload(path)
    block_name = f"full_{group}_item_macro"
    block = payload.get(block_name, {})
    if isinstance(block, dict) and all(metric in block for metric in METRICS):
        count = int(payload.get(f"count_full_{group}_item_macro", 0) or 0)
        return {metric: float(block[metric]) for metric in METRICS}, count

    values: dict[str, float] = {}
    for metric in METRICS:
        suffix = metric.lower().replace("@", "")
        column = f"full_{group}_item_macro_{suffix}"
        if column not in payload:
            raise ValueError(f"missing {block_name} metric {metric}: {path}")
        values[metric] = float(payload[column])
    count = int(payload.get(f"full_{group}_item_macro_count", 0) or 0)
    return values, count


def load_group_metrics(
    result_path: Path,
    per_item_path: Path | None,
    group: str,
) -> tuple[dict[str, float], int, str]:
    if per_item_path is not None:
        metrics, count = per_item_metrics(per_item_path)
        return metrics, count, relative(per_item_path)
    metrics, count = result_group_metrics(result_path, group)
    if count <= 0:
        raise ValueError(f"missing {group} course count: {result_path}")
    return metrics, count, relative(result_path)


def collect_core_rows() -> list[dict[str, object]]:
    warm_by_key = {
        (spec.dataset, spec.method, spec.seed): spec
        for spec in warm_export.core_specs()
    }
    rows: list[dict[str, object]] = []
    for cold_spec in cold_audit.build_specs():
        key = (cold_spec.dataset, cold_spec.method, cold_spec.seed)
        warm_spec = warm_by_key[key]
        cold_result = first_existing(cold_spec.result_candidates)
        hot_result = first_existing(warm_spec.result_candidates)
        cold_path = first_existing(cold_spec.per_item_candidates)
        hot_path = first_existing(warm_spec.per_item_candidates)
        if cold_result is None or hot_result is None:
            raise ValueError(f"missing core result source for {key}")
        cold, cold_count, cold_source = load_group_metrics(
            cold_result, cold_path, "cold"
        )
        hot, hot_count, hot_source = load_group_metrics(
            hot_result, hot_path, "hot"
        )
        rows.append(
            build_seed_row(
                dataset=cold_spec.dataset,
                method=cold_spec.method,
                seed=cold_spec.seed,
                cold=cold,
                hot=hot,
                cold_count=cold_count,
                hot_count=hot_count,
                cold_source=cold_source,
                hot_source=hot_source,
                direct=None,
            )
        )
    return rows


def collect_semco_rows() -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for spec in warm_export.semco_specs():
        result_path = first_existing(spec.result_candidates)
        hot_path = first_existing(spec.per_item_candidates)
        if result_path is None:
            raise ValueError(
                f"missing SEMCo result source for {spec.dataset}/{spec.seed}"
            )
        cold_path = None
        if hot_path is not None:
            candidate = Path(
                str(hot_path).replace(
                    "per_item_full_hot_", "per_item_full_cold_"
                )
            )
            cold_path = candidate if candidate.exists() else None
        cold, cold_count, cold_source = load_group_metrics(
            result_path, cold_path, "cold"
        )
        hot, hot_count, hot_source = load_group_metrics(
            result_path, hot_path, "hot"
        )
        rows.append(
            build_seed_row(
                dataset=spec.dataset,
                method="SEMCo",
                seed=spec.seed,
                cold=cold,
                hot=hot,
                cold_count=cold_count,
                hot_count=hot_count,
                cold_source=cold_source,
                hot_source=hot_source,
                direct=None,
            )
        )
    return rows


def collect_kgrec_rows() -> list[dict[str, object]]:
    summary_path = (
        PAPER
        / "baseline_sources/_kgrec_strict/_remaining_main_table_queue/main_table_summary.json"
    )
    summary = json.loads(summary_path.read_text(encoding="utf-8-sig"))
    rows: list[dict[str, object]] = []
    for dataset_block in summary["datasets"]:
        dataset = str(dataset_block["dataset"])
        for seed_block in dataset_block["seeds"]:
            seed = int(seed_block["seed"])
            report_path = Path(seed_block["report_path"])
            report = json.loads(report_path.read_text(encoding="utf-8-sig"))
            test = report["test"]
            counts = test["counts"]
            row = build_seed_row(
                dataset=dataset,
                method="KGRec",
                seed=seed,
                cold={metric: float(test["full_cold_item_macro"][metric]) for metric in METRICS},
                hot={metric: float(test["full_hot_item_macro"][metric]) for metric in METRICS},
                cold_count=int(counts["cold_items"]),
                hot_count=int(counts["hot_items"]),
                cold_source=relative(report_path),
                hot_source=relative(report_path),
                direct={metric: float(test["full_all_item_macro"][metric]) for metric in METRICS},
            )
            row["aggregation_route"] = "direct_all_item_macro_validated"
            rows.append(row)
    return rows


def collect_pcgnn_rows() -> list[dict[str, object]]:
    root = PAPER / "baseline_sources/_pcgnn_strict"
    prefixes = {"MOOCCube": "mooccube", "Junyi": "junyi", "COCO": "coco"}
    rows: list[dict[str, object]] = []
    for dataset, prefix in prefixes.items():
        for seed in SEEDS:
            report_path = (
                root
                / f"{prefix}_seed{seed}_full_formal_kg_warm"
                / "pcgnn_strict_adapter_report.json"
            )
            report = json.loads(report_path.read_text(encoding="utf-8-sig"))
            test = report.get("test", {})
            cold = test.get("full_cold_item_macro", {})
            cold_count = int(test.get("count_full_cold_item_macro", 0) or 0)
            if not isinstance(cold, dict) or cold_count <= 0:
                raise ValueError(f"missing PCGNN cold metrics: {report_path}")
            if float(report.get("kg_loss_weight", float("nan"))) != 1.0:
                raise ValueError(f"unexpected PCGNN kg_loss_weight: {report_path}")
            if str(report.get("rs_candidate_mode", "")) != "warm":
                raise ValueError(f"unexpected PCGNN rs_candidate_mode: {report_path}")

            row: dict[str, object] = {
                "dataset": dataset,
                "method": "PCGNN",
                "seed": int(seed),
                "status": "ready",
                "reason": (
                    "retained cold-target strict report; "
                    "no warm-target PCGNN metrics retained"
                ),
                "aggregation_route": "pcgnn_cold_target_only_report",
                "cold_count": cold_count,
                "hot_count": int(test.get("count_full_hot_item_macro", 0) or 0),
                "cold_source": relative(report_path),
                "hot_source": "",
            }
            for metric in METRICS:
                if metric not in cold:
                    raise ValueError(
                        f"missing PCGNN cold metric {metric}: {report_path}"
                    )
                row[metric] = float(cold[metric])
            rows.append(row)
    return rows


def collect_seed_rows() -> tuple[pd.DataFrame, pd.DataFrame]:
    rows = [
        *collect_core_rows(),
        *collect_semco_rows(),
        *collect_kgrec_rows(),
        *collect_pcgnn_rows(),
    ]
    coverage = pd.DataFrame(rows)
    duplicates = coverage.duplicated(["dataset", "method", "seed"], keep=False)
    if duplicates.any():
        duplicate_rows = coverage.loc[
            duplicates, ["dataset", "method", "seed"]
        ].to_dict("records")
        raise ValueError(f"duplicate dataset-method-seed rows: {duplicate_rows}")
    coverage["dataset"] = pd.Categorical(
        coverage["dataset"], categories=DATASETS, ordered=True
    )
    coverage["method"] = pd.Categorical(
        coverage["method"], categories=METHOD_ORDER, ordered=True
    )
    coverage = coverage.sort_values(["dataset", "method", "seed"]).reset_index(
        drop=True
    )
    detail = coverage.loc[coverage["status"] == "ready"].copy()
    return detail, coverage


def summarize_ready(detail: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    required_seeds = set(SEEDS)
    for (dataset, method), group in detail.groupby(
        ["dataset", "method"], observed=True, sort=False
    ):
        seeds = {int(seed) for seed in group["seed"]}
        if seeds != required_seeds:
            raise ValueError(
                f"{dataset}/{method} requires seeds 2025, 2026, 2027; "
                f"found {sorted(seeds)}"
            )
        row: dict[str, object] = {
            "dataset": str(dataset),
            "method": str(method),
            "seeds": "2025,2026,2027",
        }
        for metric in METRICS:
            values = pd.to_numeric(group[metric], errors="raise")
            row[metric] = float(values.mean())
            row[f"{metric}_std"] = float(values.std(ddof=1))
        rows.append(row)

    summary = pd.DataFrame(rows)
    for dataset in summary["dataset"].unique():
        mask = summary["dataset"] == dataset
        dataset_rows = summary.loc[mask]
        baseline_rows = dataset_rows[dataset_rows["method"] != "CKG-RL"]
        for metric in METRICS:
            summary.loc[mask, f"{metric}_rank"] = dataset_rows[metric].rank(
                method="min", ascending=False
            )
            best_index = baseline_rows[metric].idxmax()
            best_value = float(summary.loc[best_index, metric])
            best_method = str(summary.loc[best_index, "method"])
            summary.loc[mask, f"{metric}_strongest_baseline"] = best_value
            summary.loc[mask, f"{metric}_strongest_baseline_method"] = best_method
            summary.loc[mask, f"{metric}_relative_improvement"] = math.nan
            ours = mask & summary["method"].eq("CKG-RL")
            if best_value > 0 and ours.any():
                summary.loc[ours, f"{metric}_relative_improvement"] = (
                    summary.loc[ours, metric] - best_value
                ) / best_value

    for metric in METRICS:
        summary[f"{metric}_rank"] = summary[f"{metric}_rank"].astype(int)
    dataset_order = {name: index for index, name in enumerate(DATASETS)}
    method_order = {name: index for index, name in enumerate(METHOD_ORDER)}
    summary["_dataset_order"] = summary["dataset"].map(
        lambda value: dataset_order.get(str(value), len(dataset_order))
    )
    summary["_method_order"] = summary["method"].map(
        lambda value: method_order.get(str(value), len(method_order))
    )
    return summary.sort_values(
        ["_dataset_order", "_method_order", "method"]
    ).drop(columns=["_dataset_order", "_method_order"]).reset_index(drop=True)


def build_wide(summary: pd.DataFrame) -> pd.DataFrame:
    methods = [
        method
        for method in METHOD_ORDER
        if method in set(summary["method"])
    ]
    rows: list[dict[str, object]] = []
    for method in methods:
        row: dict[str, object] = {"method": method}
        for dataset in DATASETS:
            match = summary[
                summary["dataset"].eq(dataset) & summary["method"].eq(method)
            ]
            if len(match) != 1:
                raise ValueError(
                    f"expected one summary row for {dataset}/{method}, found {len(match)}"
                )
            record = match.iloc[0]
            for metric in MAIN_METRICS:
                row[f"{dataset}_{metric}"] = float(record[metric])
        rows.append(row)
    return pd.DataFrame(rows)


def plain_frame(frame: pd.DataFrame) -> pd.DataFrame:
    out = frame.copy()
    for column in out.select_dtypes(include="category").columns:
        out[column] = out[column].astype(object)
    return out


def format_worksheet(worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    body_font = Font(name="Arial", size=10)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            if isinstance(cell.value, float):
                cell.number_format = "0.0000"
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for index, column in enumerate(worksheet.iter_cols(), start=1):
        width = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column
        )
        worksheet.column_dimensions[get_column_letter(index)].width = min(
            max(width + 2, 10), 48
        )


def write_main_table_workbook(summary: pd.DataFrame, output_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Overall_Main_Table"
    sheet.sheet_view.showGridLines = False

    dark_blue = "1F4E78"
    light_blue = "D9EAF7"
    light_gray = "F2F2F2"
    border_color = "A6A6A6"
    thin = Side(style="thin", color=border_color)
    medium = Side(style="medium", color="5B5B5B")
    base_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    body_font = Font(name="Arial", size=10, color="000000")

    sheet["A1"] = "Method"
    sheet.merge_cells("A1:A2")
    for dataset_index, dataset in enumerate(DATASETS):
        start_column = 2 + dataset_index * len(MAIN_METRICS)
        end_column = start_column + len(MAIN_METRICS) - 1
        sheet.cell(1, start_column, dataset)
        sheet.merge_cells(
            start_row=1,
            start_column=start_column,
            end_row=1,
            end_column=end_column,
        )
        for metric_index, metric in enumerate(MAIN_METRICS):
            sheet.cell(2, start_column + metric_index, metric)

    for row in sheet.iter_rows(min_row=1, max_row=2, min_col=1, max_col=13):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=dark_blue)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = base_border

    records = {
        (str(row["dataset"]), str(row["method"])): row
        for _, row in summary.iterrows()
    }
    ranked_values: dict[tuple[str, str], list[float]] = {}
    for dataset in DATASETS:
        for metric in MAIN_METRICS:
            values = sorted(
                {
                    float(row[metric])
                    for _, row in summary.loc[
                        summary["dataset"].eq(dataset)
                    ].iterrows()
                },
                reverse=True,
            )
            ranked_values[(dataset, metric)] = values

    for method_index, method in enumerate(METHOD_ORDER, start=3):
        method_cell = sheet.cell(method_index, 1, method)
        method_cell.fill = PatternFill("solid", fgColor=light_gray)
        method_cell.font = Font(
            name="Arial", size=10, bold=method == "CKG-RL", color="000000"
        )
        method_cell.alignment = Alignment(horizontal="left", vertical="center")
        method_cell.border = base_border
        for dataset_index, dataset in enumerate(DATASETS):
            record = records.get((dataset, method))
            for metric_index, metric in enumerate(MAIN_METRICS):
                column = 2 + dataset_index * len(MAIN_METRICS) + metric_index
                cell = sheet.cell(method_index, column)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = base_border
                if record is None:
                    cell.value = "N/A"
                    cell.fill = PatternFill("solid", fgColor="E7E6E6")
                    cell.font = Font(
                        name="Arial", size=10, italic=True, color="7F7F7F"
                    )
                    continue
                value = float(record[metric])
                ordered = ranked_values[(dataset, metric)]
                cell.value = value
                cell.number_format = "0.0000"
                cell.font = Font(
                    name="Arial",
                    size=10,
                    bold=value == ordered[0],
                    underline="single"
                    if len(ordered) > 1 and value == ordered[1]
                    else None,
                    color="000000",
                )

    ckg_row = 3 + METHOD_ORDER.index("CKG-RL")
    imp_row = 3 + len(METHOD_ORDER)
    for column in range(1, 14):
        current = sheet.cell(ckg_row, column).border
        sheet.cell(ckg_row, column).border = Border(
            left=current.left,
            right=current.right,
            top=medium,
            bottom=current.bottom,
        )
    imp_label = sheet.cell(imp_row, 1, "Imp.")
    imp_label.fill = PatternFill("solid", fgColor=light_blue)
    imp_label.font = Font(name="Arial", size=10, bold=True, italic=True)
    imp_label.alignment = Alignment(horizontal="left", vertical="center")
    imp_label.border = Border(left=thin, right=thin, top=thin, bottom=medium)
    for dataset_index, dataset in enumerate(DATASETS):
        record = records[(dataset, "CKG-RL")]
        for metric_index, metric in enumerate(MAIN_METRICS):
            column = 2 + dataset_index * len(MAIN_METRICS) + metric_index
            cell = sheet.cell(
                imp_row,
                column,
                float(record[f"{metric}_relative_improvement"]),
            )
            cell.number_format = "+0.0%;-0.0%;0.0%"
            cell.fill = PatternFill("solid", fgColor=light_blue)
            cell.font = Font(name="Arial", size=10, italic=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=medium)

    note_row = imp_row + 2
    sheet.cell(
        note_row,
        1,
        "Note: Three-seed overall item-macro means. Bold marks best and underline marks second best. "
        "Imp. is CKG-RL relative to the strongest available baseline. Most rows reconstruct cold+hot "
        "overall values by target-course counts; PCGNN uses retained cold-target strict reports from "
        "the formal KG-warm runs because no warm-target PCGNN metrics were retained.",
    )
    sheet.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=13)
    note_cell = sheet.cell(note_row, 1)
    note_cell.font = Font(name="Arial", size=9, italic=True, color="595959")
    note_cell.alignment = Alignment(wrap_text=True, vertical="top")

    sheet.freeze_panes = "B3"
    sheet.column_dimensions["A"].width = 17
    for column in range(2, 14):
        sheet.column_dimensions[get_column_letter(column)].width = 11
    sheet.row_dimensions[1].height = 22
    sheet.row_dimensions[2].height = 20
    sheet.row_dimensions[note_row].height = 34
    sheet.auto_filter.ref = f"A2:M{2 + len(METHOD_ORDER)}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.print_title_rows = "1:2"
    sheet.print_area = f"A1:M{note_row}"
    workbook.save(output_path)


def latex_metric_value(value: float, best: float, second: float) -> str:
    if not math.isfinite(value):
        return "--"
    text = f"{value:.4f}"
    if math.isfinite(best) and abs(value - best) <= 1e-12:
        return rf"\textbf{{{text}}}"
    if math.isfinite(second) and abs(value - second) <= 1e-12:
        return rf"\underline{{{text}}}"
    return text


def build_latex_main_table_fragment(
    summary: pd.DataFrame,
    environment: str = "table*",
    placement: str = "[t]",
) -> str:
    records = {
        (str(row["dataset"]), str(row["method"])): row
        for _, row in summary.iterrows()
    }
    ranked_values: dict[tuple[str, str], tuple[float, float]] = {}
    for dataset in DATASETS:
        dataset_rows = summary.loc[summary["dataset"].eq(dataset)]
        for metric in MAIN_METRICS:
            values = sorted(
                {float(row[metric]) for _, row in dataset_rows.iterrows()},
                reverse=True,
            )
            best = values[0] if values else math.nan
            second = values[1] if len(values) > 1 else math.nan
            ranked_values[(dataset, metric)] = (best, second)

    rows: list[str] = []
    for method in METHOD_ORDER:
        cells = [METHOD_LABELS.get(method, method)]
        for dataset in DATASETS:
            record = records.get((dataset, method))
            for metric in MAIN_METRICS:
                if record is None:
                    cells.append("--")
                    continue
                value = float(record[metric])
                best, second = ranked_values[(dataset, metric)]
                cells.append(latex_metric_value(value, best, second))
        rows.append(" & ".join(cells) + r" \\")

    imp_cells = [r"\emph{Imp.}"]
    for dataset in DATASETS:
        record = records[(dataset, "CKG-RL")]
        for metric in MAIN_METRICS:
            value = float(record[f"{metric}_relative_improvement"])
            sign = "+" if value >= 0 else ""
            imp_cells.append(f"{sign}{value * 100:.1f}\\%")
    imp_row = " & ".join(imp_cells) + r" \\"

    return "\n".join(
        [
            rf"\begin{{{environment}}}{placement}",
            r"\centering",
            r"{",
            r"\small",
            r"\setlength{\tabcolsep}{1.35pt}",
            r"\renewcommand{\arraystretch}{0.98}",
            r"\begin{tabular}{lcccccccccccc}",
            r"\toprule",
            r"& \multicolumn{4}{c}{MOOCCube} & \multicolumn{4}{c}{Junyi} & \multicolumn{4}{c}{COCO} \\",
            r"\cmidrule(lr){2-5}\cmidrule(lr){6-9}\cmidrule(lr){10-13}",
            r"Method & R@5 & R@10 & N@5 & N@10 & R@5 & R@10 & N@5 & N@10 & R@5 & R@10 & N@5 & N@10 \\",
            r"\midrule",
            *rows[:-1],
            r"\midrule",
            rows[-1],
            imp_row,
            r"\bottomrule",
            r"\end{tabular}",
            r"}",
            rf"\caption{{{OVERALL_MAIN_TABLE_CAPTION}}}",
            r"\label{tab:overall-item-macro-main}",
            r"\vspace{0.2em}",
            r"\begin{minipage}{\textwidth}",
            rf"\scriptsize\emph{{Venue tags:}} {VENUE_TAGS}",
            r"\end{minipage}",
            rf"\end{{{environment}}}",
            "",
        ]
    )


def build_latex_main_table(summary: pd.DataFrame) -> str:
    return "\n".join(
        [
            r"\documentclass[10pt]{article}",
            r"\usepackage[T1]{fontenc}",
            r"\usepackage{newtxtext}",
            r"\usepackage{amsmath}",
            r"\usepackage{booktabs}",
            r"\usepackage[letterpaper,margin=0.75in]{geometry}",
            "",
            r"\pagestyle{empty}",
            r"\newcommand{\methodref}[1]{\textsuperscript{\scriptsize[#1]}}",
            "",
            r"\begin{document}",
            build_latex_main_table_fragment(
                summary, environment="table", placement="[!ht]"
            ).rstrip(),
            r"\end{document}",
            "",
        ]
    )


def write_outputs(
    detail: pd.DataFrame,
    summary: pd.DataFrame,
    coverage: pd.DataFrame,
    output_dir: Path,
) -> list[Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    detail_out = plain_frame(detail)
    summary_out = plain_frame(summary)
    coverage_out = plain_frame(coverage)
    wide = build_wide(summary_out)

    paths = {
        "workbook": output_dir / "overall_baseline_comparison.xlsx",
        "main_table": output_dir / "overall_main_table.xlsx",
        "summary": output_dir / "overall_baseline_summary.csv",
        "detail": output_dir / "overall_baseline_seed_detail.csv",
        "wide": output_dir / "overall_baseline_wide.csv",
        "coverage": output_dir / "overall_baseline_coverage.csv",
        "main_table_tex": output_dir / "overall_main_table.tex",
        "main_table_aaai_tex": output_dir / "overall_main_table_aaai.tex",
    }
    summary_out.to_csv(paths["summary"], index=False, float_format="%.15g")
    detail_out.to_csv(paths["detail"], index=False, float_format="%.15g")
    wide.to_csv(paths["wide"], index=False, float_format="%.15g")
    coverage_out.to_csv(paths["coverage"], index=False, float_format="%.15g")
    paths["main_table_tex"].write_text(
        build_latex_main_table(summary_out), encoding="ascii"
    )
    paths["main_table_aaai_tex"].write_text(
        build_latex_main_table_fragment(summary_out), encoding="ascii"
    )

    with pd.ExcelWriter(paths["workbook"], engine="openpyxl") as writer:
        summary_out.to_excel(writer, sheet_name="Summary", index=False)
        detail_out.to_excel(writer, sheet_name="Seed_Detail", index=False)
        coverage_out.to_excel(writer, sheet_name="Coverage", index=False)
        for worksheet in writer.book.worksheets:
            format_worksheet(worksheet)
    write_main_table_workbook(summary_out, paths["main_table"])
    return list(paths.values())


def main() -> None:
    detail, coverage = collect_seed_rows()
    summary = summarize_ready(detail)
    paths = write_outputs(detail, summary, coverage, OUTPUT_DIR)
    for path in paths:
        print(path)


if __name__ == "__main__":
    main()
