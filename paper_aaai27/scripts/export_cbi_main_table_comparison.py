from __future__ import annotations

import re
from pathlib import Path

import pandas as pd
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
PAPER_TEX = ROOT / "paper_aaai27" / "main.tex"
CBI_DIR = (
    ROOT
    / "outputs"
    / "cbi_faithful_single_seed2025"
    / "strict_item_cold_balanced_thr1_seed_2025"
)
CBI_REPORT = CBI_DIR / "final_report_usim_feedback_fast3_content_delta_static.csv"
CBI_LOG = ROOT / "background_logs" / "cbi_faithful_single_seed2025" / "training.log"
QUEUE_LOG = ROOT / "background_logs" / "cbi_faithful_single_seed2025" / "queue.log"
MAIN_SEED2025 = (
    ROOT
    / "outputs"
    / "significance_per_item_exports"
    / "mooccube"
    / "ckg_rl_full"
    / "strict_item_cold_balanced_thr1_seed_2025"
    / "per_item_full_cold_usim_feedback_fast3_content_delta_static.csv"
)
OUTPUT_DIR = ROOT / "paper_aaai27" / "figures" / "cbi_main_table_comparison"

METHOD_ORDER = (
    "Popularity",
    "BPR",
    "DropoutNet",
    "LightGCN",
    "CCFCRec",
    "ALDI",
    "KGRec",
    "CGRC",
    "PCGNN",
    "USIM",
    "SEMCo",
    "CKG-RL",
)
MAIN_METRICS = ("R@5", "R@10", "N@5", "N@10")
FULL_METRICS = ("R@5", "R@10", "R@20", "N@5", "N@10", "N@20")


def relative(path: Path) -> str:
    return str(path.relative_to(ROOT))


def parse_paper_main_table() -> pd.DataFrame:
    lines = PAPER_TEX.read_text(encoding="utf-8").splitlines()
    rows: list[dict[str, object]] = []
    for method in METHOD_ORDER:
        line = next(
            (
                value
                for value in lines
                if value.startswith(f"{method} &")
                or value.startswith(f"\\textbf{{{method}}} &")
            ),
            None,
        )
        if line is None:
            raise ValueError(f"missing main-table row for {method}")
        values = [float(value) for value in re.findall(r"0\.\d+", line)]
        if len(values) < len(MAIN_METRICS):
            raise ValueError(f"incomplete MOOCCube metrics for {method}")
        row: dict[str, object] = {
            "Method": method,
            "Evidence": "Paper main table (3-seed mean)",
        }
        row.update(dict(zip(MAIN_METRICS, values[:4], strict=True)))
        rows.append(row)
    return pd.DataFrame(rows)


def load_cbi_metrics() -> tuple[dict[str, float], pd.DataFrame]:
    report = pd.read_csv(CBI_REPORT).set_index("metric")
    cbi = {
        metric: float(report.loc[metric, "full_cold_item_macro"])
        for metric in FULL_METRICS
    }
    records: list[dict[str, object]] = []
    for group, column in (
        ("Cold course-macro", "full_cold_item_macro"),
        ("Hot course-macro", "full_hot_item_macro"),
    ):
        for metric in FULL_METRICS:
            records.append(
                {
                    "Group": group,
                    "Metric": metric,
                    "Value": float(report.loc[metric, column]),
                }
            )
    return cbi, pd.DataFrame(records)


def load_seed_matched_ckg() -> dict[str, float]:
    frame = pd.read_csv(MAIN_SEED2025)
    missing = [metric for metric in FULL_METRICS if metric not in frame.columns]
    if missing or frame.empty:
        raise ValueError(f"invalid seed-matched CKG-RL source: missing={missing}")
    return {
        metric: float(pd.to_numeric(frame[metric], errors="raise").mean())
        for metric in FULL_METRICS
    }


def collect_comparison() -> tuple[pd.DataFrame, dict[str, float], dict[str, float]]:
    frame = parse_paper_main_table()
    cbi, _ = load_cbi_metrics()
    seed_matched = load_seed_matched_ckg()
    cbi_row: dict[str, object] = {
        "Method": "CBI-Faithful",
        "Evidence": "Seed 2025 single run",
    }
    cbi_row.update({metric: cbi[metric] for metric in MAIN_METRICS})
    frame = pd.concat([frame, pd.DataFrame([cbi_row])], ignore_index=True)
    return frame, cbi, seed_matched


def build_delta_frame(
    cbi: dict[str, float],
    reference: dict[str, float],
    reference_name: str,
) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Metric": metric,
                "CBI-Faithful": cbi[metric],
                "Reference": reference[metric],
                "Absolute Delta": None,
                "Relative Delta": None,
                "Reference Method": reference_name,
            }
            for metric in MAIN_METRICS
        ]
    )


def provenance_frame() -> pd.DataFrame:
    log_text = CBI_LOG.read_text(encoding="utf-8", errors="replace")
    epoch_match = re.search(r"Restore best epoch=(\d+)", log_text)
    queue_text = QUEUE_LOG.read_text(encoding="utf-8", errors="replace")
    exit_match = re.search(r"CBI_EXIT code=(\d+)", queue_text)
    return pd.DataFrame(
        [
            ("Dataset", "MOOCCube"),
            ("Protocol", "Strict course-cold, full-catalog ranking"),
            ("Aggregation", "Cold course-macro"),
            ("CBI seed scope", "Single run: seed 2025"),
            ("Paper scope", "Three-seed mean: 2025, 2026, 2027"),
            ("Validation-selected epoch", epoch_match.group(1) if epoch_match else "unknown"),
            ("Queue exit code", exit_match.group(1) if exit_match else "unknown"),
            (
                "Queue status note",
                "Final report and per-item artifacts were written before the queue returned exit 1.",
            ),
            (
                "Interpretation",
                "The paper-table ranking is indicative only; confirm CBI with seeds 2026 and 2027 before changing claims.",
            ),
            ("Paper source", relative(PAPER_TEX) + ":350-368"),
            ("CBI source", relative(CBI_REPORT)),
            ("CBI training log", relative(CBI_LOG)),
            ("Seed-matched CKG-RL source", relative(MAIN_SEED2025)),
        ],
        columns=["Field", "Value"],
    )


def style_sheet(sheet, metric_columns: tuple[int, ...] = ()) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    body_font = Font(name="Arial", size=10, color="000000")
    thin = Side(style="thin", color="D9E2F3")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(bottom=thin)
    for column in metric_columns:
        for cell in sheet.iter_cols(min_col=column, max_col=column, min_row=2):
            for item in cell:
                item.number_format = "0.0000"
                item.alignment = Alignment(horizontal="right", vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    sheet.row_dimensions[1].height = 24
    for index, column_cells in enumerate(sheet.columns, start=1):
        values = [str(cell.value or "") for cell in column_cells]
        width = min(max(max(map(len, values)) + 2, 11), 70)
        sheet.column_dimensions[get_column_letter(index)].width = width


def add_delta_formulas(sheet) -> None:
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, 4, f"=B{row}-C{row}")
        sheet.cell(row, 5, f'=IF(C{row}=0,"",D{row}/C{row})')
        sheet.cell(row, 5).number_format = "0.0%"
    sheet["D1"].comment = Comment("CBI-Faithful minus reference.", "Codex")
    sheet["E1"].comment = Comment("Absolute delta divided by reference.", "Codex")


def write_aaai_table(
    frame: pd.DataFrame,
    cbi: dict[str, float],
    output_dir: Path = OUTPUT_DIR,
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / "cbi_vs_main_table_aaai.tex"
    paper_ckg = (
        frame.loc[frame["Method"] == "CKG-RL", list(MAIN_METRICS)]
        .iloc[0]
        .astype(float)
        .to_dict()
    )
    def metric(value: float, bold: bool = False, underline: bool = False) -> str:
        rendered = f"{float(value):.4f}"
        if underline:
            rendered = rf"\underline{{{rendered}}}"
        if bold:
            rendered = rf"\textbf{{{rendered}}}"
        return rendered

    lines = [
        r"\begin{table*}[t]",
        r"\centering",
        r"{",
        r"\small",
        r"\setlength{\tabcolsep}{5.5pt}",
        r"\renewcommand{\arraystretch}{0.98}",
        r"\begin{tabular*}{\textwidth}{@{\extracolsep{\fill}}lcccc}",
        r"\toprule",
        "Method & R@5 & R@10 & N@5 & N@10 " + (chr(92) * 2),
        r"\midrule",
    ]
    for _, row in frame.iterrows():
        method = str(row["Method"])
        is_cbi = method == "CBI-Faithful"
        method_label = r"\textbf{CBI-Faithful}$^\dagger$" if is_cbi else method
        underline = method == "CGRC"
        values = [
            metric(row[name], bold=is_cbi, underline=underline)
            for name in MAIN_METRICS
        ]
        lines.append(f"{method_label} & " + " & ".join(values) + " " + (chr(92) * 2))

    improvement = [
        rf"{((cbi[name] - paper_ckg[name]) / paper_ckg[name]) * 100:+.1f}\%"
        for name in MAIN_METRICS
    ]
    lines.extend(
        [
            r"\midrule",
            r"\emph{Imp. vs. CKG-RL} & " + " & ".join(improvement) + " " + (chr(92) * 2),
            r"\bottomrule",
            r"\end{tabular*}",
            r"}",
            r"\caption{Exploratory CBI-Faithful comparison on MOOCCube under strict course-cold full-ranking course-macro evaluation. Baseline and CKG-RL values are the paper's three-seed means; CBI-Faithful is a single seed 2025 run. Bold marks the CBI run and underlines mark the strongest non-CBI baseline; no significance claim is made.}",
            r"\label{tab:cbi-main-table-comparison}",
            r"\vspace{2pt}",
            r"\parbox{\textwidth}{\footnotesize $^\dagger$ CBI-Faithful uses the validation-selected seed-2025 checkpoint. The single-seed result is exploratory and should not replace the three-seed main-table evidence.}",
            r"\end{table*}",
        ]
    )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path


def write_outputs(
    frame: pd.DataFrame,
    cbi: dict[str, float],
    seed_matched: dict[str, float],
    output_dir: Path = OUTPUT_DIR,
) -> tuple[Path, Path, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    csv_path = output_dir / "cbi_vs_main_table.csv"
    xlsx_path = output_dir / "cbi_vs_main_table.xlsx"
    frame.to_csv(csv_path, index=False, encoding="utf-8-sig")

    paper_ckg = (
        frame.loc[frame["Method"] == "CKG-RL", list(MAIN_METRICS)]
        .iloc[0]
        .astype(float)
        .to_dict()
    )
    paper_cgrc = (
        frame.loc[frame["Method"] == "CGRC", list(MAIN_METRICS)]
        .iloc[0]
        .astype(float)
        .to_dict()
    )
    paper_deltas = pd.concat(
        [
            build_delta_frame(cbi, paper_ckg, "CKG-RL (paper 3-seed mean)"),
            build_delta_frame(cbi, paper_cgrc, "CGRC (strongest paper baseline)"),
        ],
        ignore_index=True,
    )
    seed_deltas = build_delta_frame(
        cbi, seed_matched, "CKG-RL (seed 2025, matched)"
    )
    _, full_cbi = load_cbi_metrics()
    provenance = provenance_frame()
    tex_path = write_aaai_table(frame, cbi, output_dir)

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        frame.to_excel(writer, sheet_name="Main_Comparison", index=False)
        paper_deltas.to_excel(writer, sheet_name="CBI_Deltas", index=False)
        seed_deltas.to_excel(writer, sheet_name="Seed_Matched", index=False)
        full_cbi.to_excel(writer, sheet_name="Full_CBI_Metrics", index=False)
        provenance.to_excel(writer, sheet_name="Provenance", index=False)

        workbook = writer.book
        main = workbook["Main_Comparison"]
        style_sheet(main, (3, 4, 5, 6))
        for row in range(2, main.max_row + 1):
            method = main.cell(row, 1).value
            if method == "CBI-Faithful":
                fill = PatternFill("solid", fgColor="E2F0D9")
            elif method == "CKG-RL":
                fill = PatternFill("solid", fgColor="D9EAF7")
            elif method == "CGRC":
                fill = PatternFill("solid", fgColor="FFF2CC")
            else:
                continue
            for cell in main[row]:
                cell.fill = fill
                cell.font = Font(name="Arial", size=10, bold=True)

        for name in ("CBI_Deltas", "Seed_Matched"):
            sheet = workbook[name]
            add_delta_formulas(sheet)
            style_sheet(sheet, (2, 3, 4))
        style_sheet(workbook["Full_CBI_Metrics"], (3,))
        style_sheet(workbook["Provenance"])
        workbook["Provenance"].column_dimensions["A"].width = 28
        workbook["Provenance"].column_dimensions["B"].width = 100

    return csv_path, xlsx_path, tex_path


def main() -> None:
    frame, cbi, seed_matched = collect_comparison()
    for path in write_outputs(frame, cbi, seed_matched):
        print(path)


if __name__ == "__main__":
    main()
