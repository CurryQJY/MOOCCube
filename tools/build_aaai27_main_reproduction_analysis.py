import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(r"D:\DeskTop\MOOCCube")
OLD = ROOT / "outputs/significance_per_item_exports/mooccube/ckg_rl_full/strict_item_cold_balanced_thr1_seed_2025"
NEW = ROOT / "outputs/significance_per_item_exports/mooccube/aaai27_main_seed2025_reproduction_v1/full_reference/strict_item_cold_balanced_thr1_seed_2025"
OUT = ROOT / "outputs/significance_per_item_exports/mooccube/aaai27_main_seed2025_reproduction_v1/seed2025_reproduction_analysis.xlsx"


def first_csv(path):
    with path.open(encoding="utf-8-sig", newline="") as f:
        return next(csv.DictReader(f))


def rows_csv(path):
    with path.open(encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def header(ws):
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def main():
    old_result = first_csv(OLD / "final_fullrank_usim_feedback_fast3_content_delta_static.csv")
    new_result = first_csv(NEW / "final_fullrank_usim_feedback_fast3_content_delta_static.csv")
    old_epochs = rows_csv(OLD / "mooc_metrics_usim_feedback_fast3_content_delta_static.csv")
    new_epochs = rows_csv(NEW / "mooc_metrics_usim_feedback_fast3_content_delta_static.csv")
    old_items = rows_csv(OLD / "per_item_full_cold_usim_feedback_fast3_content_delta_static.csv")
    new_items = rows_csv(NEW / "per_item_full_cold_usim_feedback_fast3_content_delta_static.csv")
    old_manifest = json.loads((OLD / "static_protocol_manifest.json").read_text(encoding="utf-8"))
    new_manifest = json.loads((NEW / "static_protocol_manifest.json").read_text(encoding="utf-8"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Metric Comparison"
    ws.append(["metric", "main seed2025", "reproduction", "delta", "relative delta"])
    metrics = [
        ("Cold R@5", "full_cold_item_macro_r5"),
        ("Cold R@10", "full_cold_item_macro_r10"),
        ("Cold N@5", "full_cold_item_macro_n5"),
        ("Cold N@10", "full_cold_item_macro_n10"),
        ("Hot R@5", "full_hot_item_macro_r5"),
        ("Hot R@10", "full_hot_item_macro_r10"),
        ("Hot N@5", "full_hot_item_macro_n5"),
        ("Hot N@10", "full_hot_item_macro_n10"),
    ]
    for label, key in metrics:
        row = ws.max_row + 1
        ws.append([label, float(old_result[key]), float(new_result[key])])
        ws.cell(row, 4, f"=C{row}-B{row}")
        ws.cell(row, 5, f"=D{row}/B{row}")
    header(ws)
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=4):
        for cell in row:
            cell.number_format = "0.0000"
    for cell in ws["E"][1:]:
        cell.number_format = "0.0%"
    ws.column_dimensions["A"].width = 18

    ep = wb.create_sheet("Epoch Trajectory")
    fields = ["Loss", "Val_full_cold_R@10", "Val_full_cold_N@10", "Val_full_hot_R@10", "Val_full_hot_N@10"]
    ep.append(["epoch"] + [f"old_{x}" for x in fields] + [f"new_{x}" for x in fields] + [f"delta_{x}" for x in fields])
    old_by_epoch = {int(x["Epoch"]): x for x in old_epochs}
    new_by_epoch = {int(x["Epoch"]): x for x in new_epochs}
    for epoch in sorted(set(old_by_epoch) & set(new_by_epoch)):
        old = old_by_epoch[epoch]
        new = new_by_epoch[epoch]
        row = ep.max_row + 1
        ep.append([epoch] + [float(old[x]) for x in fields] + [float(new[x]) for x in fields])
        for i in range(len(fields)):
            old_col = 2 + i
            new_col = 2 + len(fields) + i
            delta_col = 2 + 2 * len(fields) + i
            ep.cell(row, delta_col, f"={ep.cell(row,new_col).coordinate}-{ep.cell(row,old_col).coordinate}")
    header(ep)
    for col in range(2, ep.max_column + 1):
        for cell in ep.iter_cols(min_col=col, max_col=col, min_row=2):
            for c in cell:
                c.number_format = "0.0000"

    items = wb.create_sheet("Per-item Delta")
    item_metrics = ["R@5", "R@10", "N@5", "N@10"]
    items.append(["item_id", "count"] + [f"old_{x}" for x in item_metrics] + [f"new_{x}" for x in item_metrics] + [f"delta_{x}" for x in item_metrics])
    old_by_item = {int(x["item_id"]): x for x in old_items}
    new_by_item = {int(x["item_id"]): x for x in new_items}
    for item_id in sorted(set(old_by_item) & set(new_by_item)):
        old = old_by_item[item_id]
        new = new_by_item[item_id]
        row = items.max_row + 1
        items.append([item_id, int(old["count"])] + [float(old[x]) for x in item_metrics] + [float(new[x]) for x in item_metrics])
        for i in range(len(item_metrics)):
            old_col = 3 + i
            new_col = 3 + len(item_metrics) + i
            delta_col = 3 + 2 * len(item_metrics) + i
            items.cell(row, delta_col, f"={items.cell(row,new_col).coordinate}-{items.cell(row,old_col).coordinate}")
    header(items)
    for col in range(3, items.max_column + 1):
        for cell in items.iter_cols(min_col=col, max_col=col, min_row=2):
            for c in cell:
                c.number_format = "0.0000"

    audit = wb.create_sheet("Config Audit")
    audit.append(["field", "main seed2025", "reproduction", "match"])
    keys = [
        "n_epochs", "batch_size", "cold_threshold", "early_stop_average_mode",
        "early_stop_k", "early_stop_patience", "early_stop_score_mode",
        "use_content_delta", "use_pseudo_cold_train", "use_course_reward",
        "use_prereq_aux_loss", "prereq_graph_source", "feedback_course_prereq_weight",
        "feedback_course_concept_weight", "feedback_course_difficulty_weight",
        "feedback_course_redundant_weight", "feedback_course_sample_beta",
        "mask_known_pos_neg", "mask_same_item_neg", "run_sampled_eval",
    ]
    for key in keys:
        row = audit.max_row + 1
        audit.append([key, old_manifest["model_config"].get(key), new_manifest["model_config"].get(key)])
        audit.cell(row, 4, f"=B{row}=C{row}")
    audit.append(["script_sha256", old_manifest["script"]["sha256"], new_manifest["script"]["sha256"], "=B22=C22"])
    header(audit)
    audit.column_dimensions["A"].width = 38
    audit.column_dimensions["B"].width = 68
    audit.column_dimensions["C"].width = 68

    notes = wb.create_sheet("Read Me")
    for row in [
        ["Finding", "Evidence / implication"],
        ["Reproduction gate", "Failed: all four cold item-macro metrics are below the original seed 2025 result."],
        ["Best checkpoint", "Original selected epoch 60 (valid cold N@10=0.2860); reproduction selected epoch 41 (0.2598)."],
        ["Training trajectory", "Training loss is nearly identical, but validation ranking diverges from the first epochs."],
        ["Configuration", "Active configuration and split artifacts match; inactive pseudo settings also match."],
        ["Code provenance", "Original source SHA 51ea12fc... is missing; reproduction uses 207f48d6.... Source behavior drift is the leading explanation."],
        ["Decision", "Do not attach new PPO/simulator ablations to the current main-table result until a compatible Full family is established."],
    ]:
        notes.append(row)
    header(notes)
    notes.column_dimensions["A"].width = 25
    notes.column_dimensions["B"].width = 115
    for row in notes.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = Font(name="Arial", bold=cell.font.bold, color=cell.font.color)
    wb.save(OUT)
    print(OUT)


if __name__ == "__main__":
    main()
