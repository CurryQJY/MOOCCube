import csv
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(r"D:\DeskTop\MOOCCube")
BASE = ROOT / "outputs" / "recppo_research_repair"
OUT = BASE / "recppo_hparam_3seed_completion" / "recppo_hparam_analysis.xlsx"

METRICS = [
    ("cold_r5", "full_cold_item_macro_r5"),
    ("cold_r10", "full_cold_item_macro_r10"),
    ("cold_n5", "full_cold_item_macro_n5"),
    ("cold_n10", "full_cold_item_macro_n10"),
    ("hot_r5", "full_hot_item_macro_r5"),
    ("hot_r10", "full_hot_item_macro_r10"),
    ("hot_n5", "full_hot_item_macro_n5"),
    ("hot_n10", "full_hot_item_macro_n10"),
]


def read_run(sweep, weight, residual, seed, relative):
    run_dir = BASE / relative / f"strict_item_cold_balanced_thr1_seed_{seed}"
    result = run_dir / "final_fullrank_usim_feedback_fast3_content_delta_static.csv"
    with result.open(encoding="utf-8-sig", newline="") as f:
        data = next(csv.DictReader(f))
    log = run_dir / "run.log"
    raw = log.read_bytes()
    encoding = "utf-16" if raw.startswith((b"\xff\xfe", b"\xfe\xff")) else "utf-8"
    best_epoch = None
    val_score = None
    for line in raw.decode(encoding, errors="ignore").splitlines():
        if "Restore best epoch=" in line:
            best_epoch = int(line.split("epoch=")[1].split()[0])
            val_score = float(line.split("]=", 1)[1].split(")", 1)[0])
    row = [sweep, weight, residual, seed]
    row.extend(float(data[source]) for _, source in METRICS)
    row.extend([best_epoch, val_score, str(result.relative_to(ROOT))])
    return row


def collect():
    rows = []
    residual_sources = {
        (0.04, 2025): Path("recppo_residual_stage2_w050_seed2025/res004_w050"),
        (0.04, 2026): Path("final_candidate_w050_res004_seeds2026_2027"),
        (0.04, 2027): Path("final_candidate_w050_res004_seeds2026_2027"),
        (0.06, 2025): Path("weightfix_stage1_seed2025/res006_w050"),
        (0.08, 2025): Path("recppo_residual_stage2_w050_seed2025/res008_w050"),
        (0.10, 2025): Path("recppo_residual_stage2_w050_seed2025/res010_w050"),
    }
    for residual, tag in [(0.06, "w050_r006"), (0.08, "w050_r008"), (0.10, "w050_r010")]:
        for seed in (2026, 2027):
            residual_sources[(residual, seed)] = Path("recppo_hparam_3seed_completion") / tag
    for residual in (0.04, 0.06, 0.08, 0.10):
        for seed in (2025, 2026, 2027):
            rows.append(read_run("residual", 0.5, residual, seed, residual_sources[(residual, seed)]))

    weight_sources = {
        (0.5, 2025): Path("recppo_residual_stage2_w050_seed2025/res004_w050"),
        (0.5, 2026): Path("final_candidate_w050_res004_seeds2026_2027"),
        (0.5, 2027): Path("final_candidate_w050_res004_seeds2026_2027"),
        (1.0, 2025): Path("recppo_residual_stage2_seed2025/res004_w100"),
    }
    for weight, tag in [(0.25, "w025_r004"), (1.0, "w100_r004"), (1.5, "w150_r004")]:
        for seed in (2026, 2027):
            weight_sources[(weight, seed)] = Path("recppo_hparam_3seed_completion") / tag
    for weight in (0.25, 0.5, 1.0, 1.5):
        for seed in (2025, 2026, 2027):
            source = weight_sources.get((weight, seed))
            if source is not None:
                rows.append(read_run("weight", weight, 0.04, seed, source))
    return rows


def style_header(ws):
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def main():
    rows = collect()
    wb = Workbook()
    raw_ws = wb.active
    raw_ws.title = "Raw Runs"
    raw_headers = ["sweep", "weight", "residual", "seed"] + [x[0] for x in METRICS] + ["best_epoch", "valid_score", "source"]
    raw_ws.append(raw_headers)
    for row in rows:
        raw_ws.append(row)
    style_header(raw_ws)
    raw_ws.column_dimensions["O"].width = 100
    for col in range(1, 15):
        raw_ws.column_dimensions[chr(64 + col)].width = 14
    for row in raw_ws.iter_rows(min_row=2, min_col=5, max_col=12):
        for cell in row:
            cell.number_format = "0.0000"

    summary_ws = wb.create_sheet("Config Summary")
    summary_headers = ["sweep", "weight", "residual", "runs", "seeds", "complete_3seed"]
    for name, _ in METRICS:
        summary_headers.extend([f"{name}_mean", f"{name}_std"])
    summary_headers.extend(["valid_score_mean", "best_epoch_mean"])
    summary_ws.append(summary_headers)

    groups = []
    for sweep, weight, residual in [
        ("weight", 0.25, 0.04), ("weight", 0.5, 0.04), ("weight", 1.0, 0.04), ("weight", 1.5, 0.04),
        ("residual", 0.5, 0.04), ("residual", 0.5, 0.06), ("residual", 0.5, 0.08), ("residual", 0.5, 0.10),
    ]:
        indexes = [i + 2 for i, row in enumerate(rows) if row[0] == sweep and row[1] == weight and row[2] == residual]
        groups.append((sweep, weight, residual, indexes))
        excel_row = summary_ws.max_row + 1
        seeds = ",".join(str(rows[i - 2][3]) for i in indexes)
        summary_ws.append([sweep, weight, residual, len(indexes), seeds, "YES" if len(indexes) == 3 else "NO"])
        col = 7
        for raw_col in range(5, 13):
            refs = ",".join(f"'Raw Runs'!{chr(64 + raw_col)}{i}" for i in indexes)
            summary_ws.cell(excel_row, col, f"=AVERAGE({refs})")
            summary_ws.cell(excel_row, col + 1, f"=STDEV({refs})" if len(indexes) > 1 else "")
            col += 2
        refs = ",".join(f"'Raw Runs'!N{i}" for i in indexes)
        summary_ws.cell(excel_row, col, f"=AVERAGE({refs})")
        refs = ",".join(f"'Raw Runs'!M{i}" for i in indexes)
        summary_ws.cell(excel_row, col + 1, f"=AVERAGE({refs})")
    style_header(summary_ws)
    summary_ws.column_dimensions["E"].width = 20
    for col in range(7, summary_ws.max_column + 1):
        for cell in summary_ws.iter_cols(min_col=col, max_col=col, min_row=2):
            for c in cell:
                c.number_format = "0.0000"

    compare_ws = wb.create_sheet("Main Table Compare")
    compare_ws.append(["configuration", "seed_count", "R@5", "delta_vs_AAAI", "R@10", "delta_vs_AAAI", "N@5", "delta_vs_AAAI", "N@10", "delta_vs_AAAI", "selection_note"])
    compare_ws.append(["AAAI CKG-RL", 3, 0.2473, 0, 0.2863, 0, 0.1972, 0, 0.2098, 0, "published main table"])
    summary_rows = {f"{summary_ws.cell(r,1).value}:{summary_ws.cell(r,2).value}:{summary_ws.cell(r,3).value}": r for r in range(2, summary_ws.max_row + 1)}
    configs = [
        ("w=0.5,r=0.04", "residual:0.5:0.04", "current candidate; valid-selected earlier"),
        ("w=0.5,r=0.06", "residual:0.5:0.06", "best 3-seed mean valid score"),
        ("w=0.5,r=0.08", "residual:0.5:0.08", "best 3-seed cold test means; do not test-select"),
        ("w=0.25,r=0.04", "weight:0.25:0.04", "only seeds 2026/2027; seed 2025 missing"),
    ]
    for label, key, note in configs:
        source_row = summary_rows[key]
        target_row = compare_ws.max_row + 1
        compare_ws.append([label, f"='Config Summary'!D{source_row}"])
        for target_col, source_col, baseline_col in [(3, 7, 3), (5, 9, 5), (7, 11, 7), (9, 13, 9)]:
            compare_ws.cell(target_row, target_col, f"='Config Summary'!{chr(64 + source_col)}{source_row}")
            compare_ws.cell(target_row, target_col + 1, f"={chr(64 + target_col)}{target_row}-{chr(64 + baseline_col)}$2")
        compare_ws.cell(target_row, 11, note)
    style_header(compare_ws)
    compare_ws.column_dimensions["A"].width = 22
    compare_ws.column_dimensions["K"].width = 46
    for row in compare_ws.iter_rows(min_row=2, min_col=3, max_col=10):
        for cell in row:
            cell.number_format = "0.0000"

    notes_ws = wb.create_sheet("Read Me")
    notes = [
        ["Finding", "Evidence / implication"],
        ["Queue status", "12/12 newly scheduled branches completed, but this is not a complete 8-config x 3-seed grid."],
        ["Missing weight seeds", "At residual=0.04, w=0.25 and w=1.5 have only seeds 2026/2027. Run seed 2025 for both before reporting 3-seed weight sensitivity."],
        ["Validation choice", "Among the complete residual sweep, r=0.06 has the highest mean validation score (0.23217), only 0.00010 above r=0.08."],
        ["Test pattern", "r=0.08 has the strongest 3-seed cold test means, but it must not be selected solely from test results."],
        ["Weight pattern", "On the paired seeds 2026/2027, w=0.25 is strongest on all four cold metrics and validation; larger PPO update strength generally hurts cold ranking."],
        ["Epoch pattern", "Best checkpoints occur at epochs 31-33, immediately after the epoch-30 warmup. Later PPO epochs usually degrade validation."],
        ["PPO claim", "These sweeps tune PPO strength but do not prove PPO is beneficial. Warmup-only / PPO-off evaluation is still required."],
    ]
    for row in notes:
        notes_ws.append(row)
    style_header(notes_ws)
    notes_ws.column_dimensions["A"].width = 25
    notes_ws.column_dimensions["B"].width = 115
    for row in notes_ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.font = Font(name="Arial", bold=cell.font.bold, color=cell.font.color)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(OUT)


if __name__ == "__main__":
    main()
